#!usr/bin/env python
# -*- coding:utf-8 -*-
"""
@author: xinjie
@file: price_record.py
@time: 2020/03/02
"""
import openpyxl

record = [
    {
        "cpu": 1,
        "ram": 1,
        "room": "盘古",
        "disk_type": "高性能",
        "price": 1.6036359912,
        "disk_price": 0.0,
        "pay_method": "每天"
    },
    {
        "cpu": 1,
        "ram": 1,
        "room": "盘古",
        "disk_type": "高性能",
        "price": 38.0,
        "disk_price": 0.0,
        "pay_method": "月"
    },
    {
        "cpu": 1,
        "ram": 1,
        "room": "盘古",
        "disk_type": "高性能",
        "price": 387.6,
        "disk_price": 0.0,
        "pay_method": "年"
    },
    {
        "cpu": 1,
        "ram": 2,
        "room": "盘古",
        "disk_type": "高性能",
        "price": 2.9189410920000003,
        "disk_price": 0.0,
        "pay_method": "每天"
    },
    {
        "cpu": 1,
        "ram": 2,
        "room": "盘古",
        "disk_type": "高性能",
        "price": 79.6,
        "disk_price": 0.0,
        "pay_method": "月"
    },
    {
        "cpu": 1,
        "ram": 2,
        "room": "盘古",
        "disk_type": "高性能",
        "price": 811.92,
        "disk_price": 0.0,
        "pay_method": "年"
    },
    {
        "cpu": 1,
        "ram": 8,
        "room": "盘古",
        "disk_type": "高性能",
        "price": 10.810771696800002,
        "disk_price": 0.0,
        "pay_method": "每天"
    },
    {
        "cpu": 1,
        "ram": 8,
        "room": "盘古",
        "disk_type": "高性能",
        "price": 278.0,
        "disk_price": 0.0,
        "pay_method": "月"
    },
    {
        "cpu": 1,
        "ram": 8,
        "room": "盘古",
        "disk_type": "高性能",
        "price": 2835.6,
        "disk_price": 0.0,
        "pay_method": "年"
    },
    {
        "cpu": 2,
        "ram": 4,
        "room": "盘古",
        "disk_type": "高性能",
        "price": 5.837882184000001,
        "disk_price": 0.0,
        "pay_method": "每天"
    },
    {
        "cpu": 2,
        "ram": 4,
        "room": "盘古",
        "disk_type": "高性能",
        "price": 207.2,
        "disk_price": 0.0,
        "pay_method": "月"
    },
    {
        "cpu": 2,
        "ram": 4,
        "room": "盘古",
        "disk_type": "高性能",
        "price": 2113.44,
        "disk_price": 0.0,
        "pay_method": "年"
    },
    {
        "cpu": 2,
        "ram": 16,
        "room": "盘古",
        "disk_type": "高性能",
        "price": 21.621543393600003,
        "disk_price": 0.0,
        "pay_method": "每天"
    },
    {
        "cpu": 2,
        "ram": 16,
        "room": "盘古",
        "disk_type": "高性能",
        "price": 591.2,
        "disk_price": 0.0,
        "pay_method": "月"
    },
    {
        "cpu": 2,
        "ram": 16,
        "room": "盘古",
        "disk_type": "高性能",
        "price": 6030.24,
        "disk_price": 0.0,
        "pay_method": "年"
    },
    {
        "cpu": 4,
        "ram": 4,
        "room": "盘古",
        "disk_type": "高性能",
        "price": 6.4145439648,
        "disk_price": 0.0,
        "pay_method": "每天"
    },
    {
        "cpu": 4,
        "ram": 4,
        "room": "盘古",
        "disk_type": "高性能",
        "price": 304.0,
        "disk_price": 0.0,
        "pay_method": "月"
    },
    {
        "cpu": 4,
        "ram": 4,
        "room": "盘古",
        "disk_type": "高性能",
        "price": 3100.8,
        "disk_price": 0.0,
        "pay_method": "年"
    },
    {
        "cpu": 4,
        "ram": 8,
        "room": "盘古",
        "disk_type": "高性能",
        "price": 11.675764368000001,
        "disk_price": 0.0,
        "pay_method": "每天"
    },
    {
        "cpu": 4,
        "ram": 8,
        "room": "盘古",
        "disk_type": "高性能",
        "price": 432.0,
        "disk_price": 0.0,
        "pay_method": "月"
    },
    {
        "cpu": 4,
        "ram": 8,
        "room": "盘古",
        "disk_type": "高性能",
        "price": 4406.4,
        "disk_price": 0.0,
        "pay_method": "年"
    },
    {
        "cpu": 4,
        "ram": 12,
        "room": "盘古",
        "disk_type": "高性能",
        "price": 16.936984771200002,
        "disk_price": 0.0,
        "pay_method": "每天"
    },
    {
        "cpu": 4,
        "ram": 12,
        "room": "盘古",
        "disk_type": "高性能",
        "price": 560.0,
        "disk_price": 0.0,
        "pay_method": "月"
    },
    {
        "cpu": 4,
        "ram": 12,
        "room": "盘古",
        "disk_type": "高性能",
        "price": 5712.0,
        "disk_price": 0.0,
        "pay_method": "年"
    },
    {
        "cpu": 4,
        "ram": 16,
        "room": "盘古",
        "disk_type": "高性能",
        "price": 22.1982051744,
        "disk_price": 0.0,
        "pay_method": "每天"
    },
    {
        "cpu": 4,
        "ram": 16,
        "room": "盘古",
        "disk_type": "高性能",
        "price": 688.0,
        "disk_price": 0.0,
        "pay_method": "月"
    },
    {
        "cpu": 4,
        "ram": 16,
        "room": "盘古",
        "disk_type": "高性能",
        "price": 7017.6,
        "disk_price": 0.0,
        "pay_method": "年"
    },
    {
        "cpu": 4,
        "ram": 32,
        "room": "盘古",
        "disk_type": "高性能",
        "price": 43.243086787200006,
        "disk_price": 0.0,
        "pay_method": "每天"
    },
    {
        "cpu": 4,
        "ram": 32,
        "room": "盘古",
        "disk_type": "高性能",
        "price": 1200.0,
        "disk_price": 0.0,
        "pay_method": "月"
    },
    {
        "cpu": 4,
        "ram": 32,
        "room": "盘古",
        "disk_type": "高性能",
        "price": 12240.0,
        "disk_price": 0.0,
        "pay_method": "年"
    },
    {
        "cpu": 8,
        "ram": 8,
        "room": "盘古",
        "disk_type": "高性能",
        "price": 12.8290879296,
        "disk_price": 0.0,
        "pay_method": "每天"
    },
    {
        "cpu": 8,
        "ram": 8,
        "room": "盘古",
        "disk_type": "高性能",
        "price": 608.0,
        "disk_price": 0.0,
        "pay_method": "月"
    },
    {
        "cpu": 8,
        "ram": 8,
        "room": "盘古",
        "disk_type": "高性能",
        "price": 6201.6,
        "disk_price": 0.0,
        "pay_method": "年"
    },
    {
        "cpu": 8,
        "ram": 16,
        "room": "盘古",
        "disk_type": "高性能",
        "price": 23.351528736000002,
        "disk_price": 0.0,
        "pay_method": "每天"
    },
    {
        "cpu": 8,
        "ram": 16,
        "room": "盘古",
        "disk_type": "高性能",
        "price": 864.0,
        "disk_price": 0.0,
        "pay_method": "月"
    },
    {
        "cpu": 8,
        "ram": 16,
        "room": "盘古",
        "disk_type": "高性能",
        "price": 8812.8,
        "disk_price": 0.0,
        "pay_method": "年"
    },
    {
        "cpu": 8,
        "ram": 24,
        "room": "盘古",
        "disk_type": "高性能",
        "price": 33.873969542400005,
        "disk_price": 0.0,
        "pay_method": "每天"
    },
    {
        "cpu": 8,
        "ram": 24,
        "room": "盘古",
        "disk_type": "高性能",
        "price": 1120.0,
        "disk_price": 0.0,
        "pay_method": "月"
    },
    {
        "cpu": 8,
        "ram": 24,
        "room": "盘古",
        "disk_type": "高性能",
        "price": 11424.0,
        "disk_price": 0.0,
        "pay_method": "年"
    },
    {
        "cpu": 8,
        "ram": 32,
        "room": "盘古",
        "disk_type": "高性能",
        "price": 44.3964103488,
        "disk_price": 0.0,
        "pay_method": "每天"
    },
    {
        "cpu": 8,
        "ram": 32,
        "room": "盘古",
        "disk_type": "高性能",
        "price": 1376.0,
        "disk_price": 0.0,
        "pay_method": "月"
    },
    {
        "cpu": 8,
        "ram": 32,
        "room": "盘古",
        "disk_type": "高性能",
        "price": 14035.2,
        "disk_price": 0.0,
        "pay_method": "年"
    },
    {
        "cpu": 8,
        "ram": 48,
        "room": "盘古",
        "disk_type": "高性能",
        "price": 65.44129196160002,
        "disk_price": 0.0,
        "pay_method": "每天"
    },
    {
        "cpu": 8,
        "ram": 48,
        "room": "盘古",
        "disk_type": "高性能",
        "price": 1888.0,
        "disk_price": 0.0,
        "pay_method": "月"
    },
    {
        "cpu": 8,
        "ram": 48,
        "room": "盘古",
        "disk_type": "高性能",
        "price": 19257.6,
        "disk_price": 0.0,
        "pay_method": "年"
    },
    {
        "cpu": 8,
        "ram": 64,
        "room": "盘古",
        "disk_type": "高性能",
        "price": 86.48617357440001,
        "disk_price": 0.0,
        "pay_method": "每天"
    },
    {
        "cpu": 8,
        "ram": 64,
        "room": "盘古",
        "disk_type": "高性能",
        "price": 2400.0,
        "disk_price": 0.0,
        "pay_method": "月"
    },
    {
        "cpu": 8,
        "ram": 64,
        "room": "盘古",
        "disk_type": "高性能",
        "price": 24480.0,
        "disk_price": 0.0,
        "pay_method": "年"
    },
    {
        "cpu": 8,
        "ram": 96,
        "room": "盘古",
        "disk_type": "高性能",
        "price": 128.57593680000002,
        "disk_price": 0.0,
        "pay_method": "每天"
    },
    {
        "cpu": 8,
        "ram": 96,
        "room": "盘古",
        "disk_type": "高性能",
        "price": 3731.2,
        "disk_price": 0.0,
        "pay_method": "月"
    },
    {
        "cpu": 8,
        "ram": 96,
        "room": "盘古",
        "disk_type": "高性能",
        "price": 38058.24,
        "disk_price": 0.0,
        "pay_method": "年"
    },
    {
        "cpu": 8,
        "ram": 128,
        "room": "盘古",
        "disk_type": "高性能",
        "price": 170.6657000256,
        "disk_price": 0.0,
        "pay_method": "每天"
    },
    {
        "cpu": 8,
        "ram": 128,
        "room": "盘古",
        "disk_type": "高性能",
        "price": 4857.6,
        "disk_price": 0.0,
        "pay_method": "月"
    },
    {
        "cpu": 8,
        "ram": 128,
        "room": "盘古",
        "disk_type": "高性能",
        "price": 49547.52,
        "disk_price": 0.0,
        "pay_method": "年"
    },
    {
        "cpu": 10,
        "ram": 16,
        "room": "盘古",
        "disk_type": "高性能",
        "price": 23.928190516800004,
        "disk_price": 0.0,
        "pay_method": "每天"
    },
    {
        "cpu": 10,
        "ram": 16,
        "room": "盘古",
        "disk_type": "高性能",
        "price": 952.0,
        "disk_price": 0.0,
        "pay_method": "月"
    },
    {
        "cpu": 10,
        "ram": 16,
        "room": "盘古",
        "disk_type": "高性能",
        "price": 9710.4,
        "disk_price": 0.0,
        "pay_method": "年"
    },
    {
        "cpu": 10,
        "ram": 24,
        "room": "盘古",
        "disk_type": "高性能",
        "price": 34.45063132320001,
        "disk_price": 0.0,
        "pay_method": "每天"
    },
    {
        "cpu": 10,
        "ram": 24,
        "room": "盘古",
        "disk_type": "高性能",
        "price": 1208.0,
        "disk_price": 0.0,
        "pay_method": "月"
    },
    {
        "cpu": 10,
        "ram": 24,
        "room": "盘古",
        "disk_type": "高性能",
        "price": 12321.6,
        "disk_price": 0.0,
        "pay_method": "年"
    },
    {
        "cpu": 10,
        "ram": 32,
        "room": "盘古",
        "disk_type": "高性能",
        "price": 44.9730721296,
        "disk_price": 0.0,
        "pay_method": "每天"
    },
    {
        "cpu": 10,
        "ram": 32,
        "room": "盘古",
        "disk_type": "高性能",
        "price": 1464.0,
        "disk_price": 0.0,
        "pay_method": "月"
    },
    {
        "cpu": 10,
        "ram": 32,
        "room": "盘古",
        "disk_type": "高性能",
        "price": 14932.8,
        "disk_price": 0.0,
        "pay_method": "年"
    },
    {
        "cpu": 10,
        "ram": 48,
        "room": "盘古",
        "disk_type": "高性能",
        "price": 66.01795374240001,
        "disk_price": 0.0,
        "pay_method": "每天"
    },
    {
        "cpu": 10,
        "ram": 48,
        "room": "盘古",
        "disk_type": "高性能",
        "price": 1976.0,
        "disk_price": 0.0,
        "pay_method": "月"
    },
    {
        "cpu": 10,
        "ram": 48,
        "room": "盘古",
        "disk_type": "高性能",
        "price": 20155.2,
        "disk_price": 0.0,
        "pay_method": "年"
    },
    {
        "cpu": 10,
        "ram": 64,
        "room": "盘古",
        "disk_type": "高性能",
        "price": 87.06283535520001,
        "disk_price": 0.0,
        "pay_method": "每天"
    },
    {
        "cpu": 10,
        "ram": 64,
        "room": "盘古",
        "disk_type": "高性能",
        "price": 2488.0,
        "disk_price": 0.0,
        "pay_method": "月"
    },
    {
        "cpu": 10,
        "ram": 64,
        "room": "盘古",
        "disk_type": "高性能",
        "price": 25377.6,
        "disk_price": 0.0,
        "pay_method": "年"
    },
    {
        "cpu": 10,
        "ram": 96,
        "room": "盘古",
        "disk_type": "高性能",
        "price": 129.1525985808,
        "disk_price": 0.0,
        "pay_method": "每天"
    },
    {
        "cpu": 10,
        "ram": 96,
        "room": "盘古",
        "disk_type": "高性能",
        "price": 3819.2,
        "disk_price": 0.0,
        "pay_method": "月"
    },
    {
        "cpu": 10,
        "ram": 96,
        "room": "盘古",
        "disk_type": "高性能",
        "price": 38955.84,
        "disk_price": 0.0,
        "pay_method": "年"
    },
    {
        "cpu": 10,
        "ram": 128,
        "room": "盘古",
        "disk_type": "高性能",
        "price": 171.2423618064,
        "disk_price": 0.0,
        "pay_method": "每天"
    },
    {
        "cpu": 10,
        "ram": 128,
        "room": "盘古",
        "disk_type": "高性能",
        "price": 4945.6,
        "disk_price": 0.0,
        "pay_method": "月"
    },
    {
        "cpu": 10,
        "ram": 128,
        "room": "盘古",
        "disk_type": "高性能",
        "price": 50445.12,
        "disk_price": 0.0,
        "pay_method": "年"
    },
    {
        "cpu": 12,
        "ram": 24,
        "room": "盘古",
        "disk_type": "高性能",
        "price": 35.02729310400001,
        "disk_price": 0.0,
        "pay_method": "每天"
    },
    {
        "cpu": 12,
        "ram": 24,
        "room": "盘古",
        "disk_type": "高性能",
        "price": 1348.8,
        "disk_price": 0.0,
        "pay_method": "月"
    },
    {
        "cpu": 12,
        "ram": 24,
        "room": "盘古",
        "disk_type": "高性能",
        "price": 13757.76,
        "disk_price": 0.0,
        "pay_method": "年"
    },
    {
        "cpu": 16,
        "ram": 16,
        "room": "盘古",
        "disk_type": "高性能",
        "price": 25.6581758592,
        "disk_price": 0.0,
        "pay_method": "每天"
    },
    {
        "cpu": 16,
        "ram": 16,
        "room": "盘古",
        "disk_type": "高性能",
        "price": 1286.4,
        "disk_price": 0.0,
        "pay_method": "月"
    },
    {
        "cpu": 16,
        "ram": 16,
        "room": "盘古",
        "disk_type": "高性能",
        "price": 13121.28,
        "disk_price": 0.0,
        "pay_method": "年"
    },
    {
        "cpu": 32,
        "ram": 32,
        "room": "盘古",
        "disk_type": "高性能",
        "price": 51.3163517184,
        "disk_price": 0.0,
        "pay_method": "每天"
    },
    {
        "cpu": 32,
        "ram": 32,
        "room": "盘古",
        "disk_type": "高性能",
        "price": 2572.8,
        "disk_price": 0.0,
        "pay_method": "月"
    },
    {
        "cpu": 32,
        "ram": 32,
        "room": "盘古",
        "disk_type": "高性能",
        "price": 26242.56,
        "disk_price": 0.0,
        "pay_method": "年"
    },
    {
        "cpu": 2,
        "ram": 2,
        "room": "盘古",
        "disk_type": "高性能",
        "price": 3.2072719824,
        "disk_price": 0.0,
        "pay_method": "每天"
    },
    {
        "cpu": 2,
        "ram": 2,
        "room": "盘古",
        "disk_type": "高性能",
        "price": 136.8,
        "disk_price": 0.0,
        "pay_method": "月"
    },
    {
        "cpu": 2,
        "ram": 2,
        "room": "盘古",
        "disk_type": "高性能",
        "price": 1395.36,
        "disk_price": 0.0,
        "pay_method": "年"
    },
    {
        "cpu": 4,
        "ram": 16,
        "room": "盘古",
        "disk_type": "高IO",
        "price": 16.5862944,
        "disk_price": 0.425088,
        "pay_method": "每天"
    },
    {
        "cpu": 4,
        "ram": 16,
        "room": "盘古",
        "disk_type": "高IO",
        "price": 613.68,
        "disk_price": 60.0,
        "pay_method": "月"
    },
    {
        "cpu": 4,
        "ram": 16,
        "room": "盘古",
        "disk_type": "高IO",
        "price": 6258.8,
        "disk_price": 720.0,
        "pay_method": "年"
    },
    {
        "cpu": 8,
        "ram": 32,
        "room": "盘古",
        "disk_type": "高IO",
        "price": 33.1725888,
        "disk_price": 0.425088,
        "pay_method": "每天"
    },
    {
        "cpu": 8,
        "ram": 32,
        "room": "盘古",
        "disk_type": "高IO",
        "price": 1227.36,
        "disk_price": 60.0,
        "pay_method": "月"
    },
    {
        "cpu": 8,
        "ram": 32,
        "room": "盘古",
        "disk_type": "高IO",
        "price": 12517.6,
        "disk_price": 720.0,
        "pay_method": "年"
    },
    {
        "cpu": 16,
        "ram": 64,
        "room": "盘古",
        "disk_type": "高IO",
        "price": 66.3451776,
        "disk_price": 0.425088,
        "pay_method": "每天"
    },
    {
        "cpu": 16,
        "ram": 64,
        "room": "盘古",
        "disk_type": "高IO",
        "price": 2454.72,
        "disk_price": 60.0,
        "pay_method": "月"
    },
    {
        "cpu": 16,
        "ram": 64,
        "room": "盘古",
        "disk_type": "高IO",
        "price": 25035.2,
        "disk_price": 720.0,
        "pay_method": "年"
    },
    {
        "cpu": 32,
        "ram": 128,
        "room": "盘古",
        "disk_type": "高IO",
        "price": 132.6903552,
        "disk_price": 0.425088,
        "pay_method": "每天"
    },
    {
        "cpu": 32,
        "ram": 128,
        "room": "盘古",
        "disk_type": "高IO",
        "price": 4909.44,
        "disk_price": 60.0,
        "pay_method": "月"
    },
    {
        "cpu": 32,
        "ram": 128,
        "room": "盘古",
        "disk_type": "高IO",
        "price": 50070.4,
        "disk_price": 720.0,
        "pay_method": "年"
    },
    {
        "cpu": "1",
        "ram": "1",
        "room": "兆维",
        "disk_type": "标准型",
        "price": 1.4587582075200003,
        "disk_price": 0.0,
        "pay_method": "每天"
    },
    {
        "cpu": "1",
        "ram": "1",
        "room": "兆维",
        "disk_type": "标准型",
        "price": 27.0,
        "disk_price": 0.0,
        "pay_method": "月"
    },
    {
        "cpu": "1",
        "ram": "1",
        "room": "兆维",
        "disk_type": "标准型",
        "price": 275.4,
        "disk_price": 0.0,
        "pay_method": "年"
    },
    {
        "cpu": "1",
        "ram": "2",
        "room": "兆维",
        "disk_type": "标准型",
        "price": 2.7738208238400004,
        "disk_price": 0.0,
        "pay_method": "每天"
    },
    {
        "cpu": "1",
        "ram": "2",
        "room": "兆维",
        "disk_type": "标准型",
        "price": 68.6,
        "disk_price": 0.0,
        "pay_method": "月"
    },
    {
        "cpu": "1",
        "ram": "2",
        "room": "兆维",
        "disk_type": "标准型",
        "price": 699.72,
        "disk_price": 0.0,
        "pay_method": "年"
    },
    {
        "cpu": "1",
        "ram": "4",
        "room": "兆维",
        "disk_type": "标准型",
        "price": 5.4039460564800015,
        "disk_price": 0.0,
        "pay_method": "每天"
    },
    {
        "cpu": "1",
        "ram": "4",
        "room": "兆维",
        "disk_type": "标准型",
        "price": 139.0,
        "disk_price": 0.0,
        "pay_method": "月"
    },
    {
        "cpu": "1",
        "ram": "4",
        "room": "兆维",
        "disk_type": "标准型",
        "price": 1417.8,
        "disk_price": 0.0,
        "pay_method": "年"
    },
    {
        "cpu": "1",
        "ram": "8",
        "room": "兆维",
        "disk_type": "标准型",
        "price": 10.664196521760001,
        "disk_price": 0.0,
        "pay_method": "每天"
    },
    {
        "cpu": "1",
        "ram": "8",
        "room": "兆维",
        "disk_type": "标准型",
        "price": 267.0,
        "disk_price": 0.0,
        "pay_method": "月"
    },
    {
        "cpu": "1",
        "ram": "8",
        "room": "兆维",
        "disk_type": "标准型",
        "price": 2723.4,
        "disk_price": 0.0,
        "pay_method": "年"
    },
    {
        "cpu": "2",
        "ram": "2",
        "room": "兆维",
        "disk_type": "标准型",
        "price": 2.9175164150400006,
        "disk_price": 0.0,
        "pay_method": "每天"
    },
    {
        "cpu": "2",
        "ram": "2",
        "room": "兆维",
        "disk_type": "标准型",
        "price": 97.2,
        "disk_price": 0.0,
        "pay_method": "月"
    },
    {
        "cpu": "2",
        "ram": "2",
        "room": "兆维",
        "disk_type": "标准型",
        "price": 991.44,
        "disk_price": 0.0,
        "pay_method": "年"
    },
    {
        "cpu": "2",
        "ram": "4",
        "room": "兆维",
        "disk_type": "标准型",
        "price": 5.547641647680001,
        "disk_price": 0.0,
        "pay_method": "每天"
    },
    {
        "cpu": "2",
        "ram": "4",
        "room": "兆维",
        "disk_type": "标准型",
        "price": 167.6,
        "disk_price": 0.0,
        "pay_method": "月"
    },
    {
        "cpu": "2",
        "ram": "4",
        "room": "兆维",
        "disk_type": "标准型",
        "price": 1709.52,
        "disk_price": 0.0,
        "pay_method": "年"
    },
    {
        "cpu": "2",
        "ram": "8",
        "room": "兆维",
        "disk_type": "标准型",
        "price": 10.807892112960003,
        "disk_price": 0.0,
        "pay_method": "每天"
    },
    {
        "cpu": "2",
        "ram": "8",
        "room": "兆维",
        "disk_type": "标准型",
        "price": 295.6,
        "disk_price": 0.0,
        "pay_method": "月"
    },
    {
        "cpu": "2",
        "ram": "8",
        "room": "兆维",
        "disk_type": "标准型",
        "price": 3015.12,
        "disk_price": 0.0,
        "pay_method": "年"
    },
    {
        "cpu": "2",
        "ram": "16",
        "room": "兆维",
        "disk_type": "标准型",
        "price": 21.328393043520002,
        "disk_price": 0.0,
        "pay_method": "每天"
    },
    {
        "cpu": "2",
        "ram": "16",
        "room": "兆维",
        "disk_type": "标准型",
        "price": 551.6,
        "disk_price": 0.0,
        "pay_method": "月"
    },
    {
        "cpu": "2",
        "ram": "16",
        "room": "兆维",
        "disk_type": "标准型",
        "price": 5626.32,
        "disk_price": 0.0,
        "pay_method": "年"
    },
    {
        "cpu": "4",
        "ram": "4",
        "room": "兆维",
        "disk_type": "标准型",
        "price": 5.835032830080001,
        "disk_price": 0.0,
        "pay_method": "每天"
    },
    {
        "cpu": "4",
        "ram": "4",
        "room": "兆维",
        "disk_type": "标准型",
        "price": 216.0,
        "disk_price": 0.0,
        "pay_method": "月"
    },
    {
        "cpu": "4",
        "ram": "4",
        "room": "兆维",
        "disk_type": "标准型",
        "price": 2203.2,
        "disk_price": 0.0,
        "pay_method": "年"
    },
    {
        "cpu": "4",
        "ram": "12",
        "room": "兆维",
        "disk_type": "标准型",
        "price": 16.355533760640004,
        "disk_price": 0.0,
        "pay_method": "每天"
    },
    {
        "cpu": "4",
        "ram": "12",
        "room": "兆维",
        "disk_type": "标准型",
        "price": 472.0,
        "disk_price": 0.0,
        "pay_method": "月"
    },
    {
        "cpu": "4",
        "ram": "12",
        "room": "兆维",
        "disk_type": "标准型",
        "price": 4814.4,
        "disk_price": 0.0,
        "pay_method": "年"
    },
    {
        "cpu": "4",
        "ram": "16",
        "room": "兆维",
        "disk_type": "标准型",
        "price": 21.615784225920006,
        "disk_price": 0.0,
        "pay_method": "每天"
    },
    {
        "cpu": "4",
        "ram": "16",
        "room": "兆维",
        "disk_type": "标准型",
        "price": 600.0,
        "disk_price": 0.0,
        "pay_method": "月"
    },
    {
        "cpu": "4",
        "ram": "16",
        "room": "兆维",
        "disk_type": "标准型",
        "price": 6120.0,
        "disk_price": 0.0,
        "pay_method": "年"
    },
    {
        "cpu": "4",
        "ram": "32",
        "room": "兆维",
        "disk_type": "标准型",
        "price": 42.656786087040004,
        "disk_price": 0.0,
        "pay_method": "每天"
    },
    {
        "cpu": "4",
        "ram": "32",
        "room": "兆维",
        "disk_type": "标准型",
        "price": 1112.0,
        "disk_price": 0.0,
        "pay_method": "月"
    },
    {
        "cpu": "4",
        "ram": "32",
        "room": "兆维",
        "disk_type": "标准型",
        "price": 11342.4,
        "disk_price": 0.0,
        "pay_method": "年"
    },
    {
        "cpu": "8",
        "ram": "8",
        "room": "兆维",
        "disk_type": "标准型",
        "price": 11.670065660160002,
        "disk_price": 0.0,
        "pay_method": "每天"
    },
    {
        "cpu": "8",
        "ram": "8",
        "room": "兆维",
        "disk_type": "标准型",
        "price": 432.0,
        "disk_price": 0.0,
        "pay_method": "月"
    },
    {
        "cpu": "8",
        "ram": "8",
        "room": "兆维",
        "disk_type": "标准型",
        "price": 4406.4,
        "disk_price": 0.0,
        "pay_method": "年"
    },
    {
        "cpu": "8",
        "ram": "16",
        "room": "兆维",
        "disk_type": "标准型",
        "price": 22.190566590720003,
        "disk_price": 0.0,
        "pay_method": "每天"
    },
    {
        "cpu": "8",
        "ram": "16",
        "room": "兆维",
        "disk_type": "标准型",
        "price": 688.0,
        "disk_price": 0.0,
        "pay_method": "月"
    },
    {
        "cpu": "8",
        "ram": "16",
        "room": "兆维",
        "disk_type": "标准型",
        "price": 7017.6,
        "disk_price": 0.0,
        "pay_method": "年"
    },
    {
        "cpu": "8",
        "ram": "24",
        "room": "兆维",
        "disk_type": "标准型",
        "price": 32.71106752128001,
        "disk_price": 0.0,
        "pay_method": "每天"
    },
    {
        "cpu": "8",
        "ram": "24",
        "room": "兆维",
        "disk_type": "标准型",
        "price": 944.0,
        "disk_price": 0.0,
        "pay_method": "月"
    },
    {
        "cpu": "8",
        "ram": "24",
        "room": "兆维",
        "disk_type": "标准型",
        "price": 9628.8,
        "disk_price": 0.0,
        "pay_method": "年"
    },
    {
        "cpu": "8",
        "ram": "32",
        "room": "兆维",
        "disk_type": "标准型",
        "price": 43.23156845184001,
        "disk_price": 0.0,
        "pay_method": "每天"
    },
    {
        "cpu": "8",
        "ram": "32",
        "room": "兆维",
        "disk_type": "标准型",
        "price": 1200.0,
        "disk_price": 0.0,
        "pay_method": "月"
    },
    {
        "cpu": "8",
        "ram": "32",
        "room": "兆维",
        "disk_type": "标准型",
        "price": 12240.0,
        "disk_price": 0.0,
        "pay_method": "年"
    },
    {
        "cpu": "8",
        "ram": "48",
        "room": "兆维",
        "disk_type": "标准型",
        "price": 64.27257031296,
        "disk_price": 0.0,
        "pay_method": "每天"
    },
    {
        "cpu": "8",
        "ram": "48",
        "room": "兆维",
        "disk_type": "标准型",
        "price": 1712.0,
        "disk_price": 0.0,
        "pay_method": "月"
    },
    {
        "cpu": "8",
        "ram": "48",
        "room": "兆维",
        "disk_type": "标准型",
        "price": 17462.4,
        "disk_price": 0.0,
        "pay_method": "年"
    },
    {
        "cpu": "8",
        "ram": "64",
        "room": "兆维",
        "disk_type": "标准型",
        "price": 85.31357217408001,
        "disk_price": 0.0,
        "pay_method": "每天"
    },
    {
        "cpu": "8",
        "ram": "64",
        "room": "兆维",
        "disk_type": "标准型",
        "price": 2224.0,
        "disk_price": 0.0,
        "pay_method": "月"
    },
    {
        "cpu": "8",
        "ram": "64",
        "room": "兆维",
        "disk_type": "标准型",
        "price": 22684.8,
        "disk_price": 0.0,
        "pay_method": "年"
    },
    {
        "cpu": "8",
        "ram": "96",
        "room": "兆维",
        "disk_type": "标准型",
        "price": 127.39557589632001,
        "disk_price": 0.0,
        "pay_method": "每天"
    },
    {
        "cpu": "8",
        "ram": "96",
        "room": "兆维",
        "disk_type": "标准型",
        "price": 3555.2,
        "disk_price": 0.0,
        "pay_method": "月"
    },
    {
        "cpu": "8",
        "ram": "96",
        "room": "兆维",
        "disk_type": "标准型",
        "price": 36263.04,
        "disk_price": 0.0,
        "pay_method": "年"
    },
    {
        "cpu": "8",
        "ram": "128",
        "room": "兆维",
        "disk_type": "标准型",
        "price": 169.47757961856001,
        "disk_price": 0.0,
        "pay_method": "每天"
    },
    {
        "cpu": "8",
        "ram": "128",
        "room": "兆维",
        "disk_type": "标准型",
        "price": 4681.6,
        "disk_price": 0.0,
        "pay_method": "月"
    },
    {
        "cpu": "8",
        "ram": "128",
        "room": "兆维",
        "disk_type": "标准型",
        "price": 47752.32,
        "disk_price": 0.0,
        "pay_method": "年"
    },
    {
        "cpu": "10",
        "ram": "16",
        "room": "兆维",
        "disk_type": "标准型",
        "price": 22.477957773120004,
        "disk_price": 0.0,
        "pay_method": "每天"
    },
    {
        "cpu": "10",
        "ram": "16",
        "room": "兆维",
        "disk_type": "标准型",
        "price": 732.0,
        "disk_price": 0.0,
        "pay_method": "月"
    },
    {
        "cpu": "10",
        "ram": "16",
        "room": "兆维",
        "disk_type": "标准型",
        "price": 7466.4,
        "disk_price": 0.0,
        "pay_method": "年"
    },
    {
        "cpu": "10",
        "ram": "24",
        "room": "兆维",
        "disk_type": "标准型",
        "price": 32.99845870368001,
        "disk_price": 0.0,
        "pay_method": "每天"
    },
    {
        "cpu": "10",
        "ram": "24",
        "room": "兆维",
        "disk_type": "标准型",
        "price": 988.0,
        "disk_price": 0.0,
        "pay_method": "月"
    },
    {
        "cpu": "10",
        "ram": "24",
        "room": "兆维",
        "disk_type": "标准型",
        "price": 10077.6,
        "disk_price": 0.0,
        "pay_method": "年"
    },
    {
        "cpu": "10",
        "ram": "32",
        "room": "兆维",
        "disk_type": "标准型",
        "price": 43.518959634240005,
        "disk_price": 0.0,
        "pay_method": "每天"
    },
    {
        "cpu": "10",
        "ram": "32",
        "room": "兆维",
        "disk_type": "标准型",
        "price": 1244.0,
        "disk_price": 0.0,
        "pay_method": "月"
    },
    {
        "cpu": "10",
        "ram": "32",
        "room": "兆维",
        "disk_type": "标准型",
        "price": 12688.8,
        "disk_price": 0.0,
        "pay_method": "年"
    },
    {
        "cpu": "10",
        "ram": "48",
        "room": "兆维",
        "disk_type": "标准型",
        "price": 64.55996149536,
        "disk_price": 0.0,
        "pay_method": "每天"
    },
    {
        "cpu": "10",
        "ram": "48",
        "room": "兆维",
        "disk_type": "标准型",
        "price": 1756.0,
        "disk_price": 0.0,
        "pay_method": "月"
    },
    {
        "cpu": "10",
        "ram": "48",
        "room": "兆维",
        "disk_type": "标准型",
        "price": 17911.2,
        "disk_price": 0.0,
        "pay_method": "年"
    },
    {
        "cpu": "10",
        "ram": "64",
        "room": "兆维",
        "disk_type": "标准型",
        "price": 85.60096335648001,
        "disk_price": 0.0,
        "pay_method": "每天"
    },
    {
        "cpu": "10",
        "ram": "64",
        "room": "兆维",
        "disk_type": "标准型",
        "price": 2268.0,
        "disk_price": 0.0,
        "pay_method": "月"
    },
    {
        "cpu": "10",
        "ram": "64",
        "room": "兆维",
        "disk_type": "标准型",
        "price": 23133.6,
        "disk_price": 0.0,
        "pay_method": "年"
    },
    {
        "cpu": "10",
        "ram": "96",
        "room": "兆维",
        "disk_type": "标准型",
        "price": 127.68296707872003,
        "disk_price": 0.0,
        "pay_method": "每天"
    },
    {
        "cpu": "10",
        "ram": "96",
        "room": "兆维",
        "disk_type": "标准型",
        "price": 3599.2,
        "disk_price": 0.0,
        "pay_method": "月"
    },
    {
        "cpu": "10",
        "ram": "96",
        "room": "兆维",
        "disk_type": "标准型",
        "price": 36711.84,
        "disk_price": 0.0,
        "pay_method": "年"
    },
    {
        "cpu": "10",
        "ram": "128",
        "room": "兆维",
        "disk_type": "标准型",
        "price": 169.76497080096004,
        "disk_price": 0.0,
        "pay_method": "每天"
    },
    {
        "cpu": "10",
        "ram": "128",
        "room": "兆维",
        "disk_type": "标准型",
        "price": 4725.6,
        "disk_price": 0.0,
        "pay_method": "月"
    },
    {
        "cpu": "10",
        "ram": "128",
        "room": "兆维",
        "disk_type": "标准型",
        "price": 48201.12,
        "disk_price": 0.0,
        "pay_method": "年"
    },
    {
        "cpu": "4",
        "ram": "8",
        "room": "兆维",
        "disk_type": "标准型",
        "price": 11.095283295360002,
        "disk_price": 0.0,
        "pay_method": "每天"
    },
    {
        "cpu": "4",
        "ram": "8",
        "room": "兆维",
        "disk_type": "标准型",
        "price": 344.0,
        "disk_price": 0.0,
        "pay_method": "月"
    },
    {
        "cpu": "4",
        "ram": "8",
        "room": "兆维",
        "disk_type": "标准型",
        "price": 3508.8,
        "disk_price": 0.0,
        "pay_method": "年"
    },
    {
        "cpu": "1",
        "ram": "1",
        "room": "兆维",
        "disk_type": "高性能",
        "price": 1.6036359912,
        "disk_price": 0.0,
        "pay_method": "每天"
    },
    {
        "cpu": "1",
        "ram": "1",
        "room": "兆维",
        "disk_type": "高性能",
        "price": 38.0,
        "disk_price": 0.0,
        "pay_method": "月"
    },
    {
        "cpu": "1",
        "ram": "1",
        "room": "兆维",
        "disk_type": "高性能",
        "price": 387.6,
        "disk_price": 0.0,
        "pay_method": "年"
    },
    {
        "cpu": "1",
        "ram": "2",
        "room": "兆维",
        "disk_type": "高性能",
        "price": 2.9189410920000003,
        "disk_price": 0.0,
        "pay_method": "每天"
    },
    {
        "cpu": "1",
        "ram": "2",
        "room": "兆维",
        "disk_type": "高性能",
        "price": 79.6,
        "disk_price": 0.0,
        "pay_method": "月"
    },
    {
        "cpu": "1",
        "ram": "2",
        "room": "兆维",
        "disk_type": "高性能",
        "price": 811.92,
        "disk_price": 0.0,
        "pay_method": "年"
    },
    {
        "cpu": "1",
        "ram": "8",
        "room": "兆维",
        "disk_type": "高性能",
        "price": 10.810771696800002,
        "disk_price": 0.0,
        "pay_method": "每天"
    },
    {
        "cpu": "1",
        "ram": "8",
        "room": "兆维",
        "disk_type": "高性能",
        "price": 278.0,
        "disk_price": 0.0,
        "pay_method": "月"
    },
    {
        "cpu": "1",
        "ram": "8",
        "room": "兆维",
        "disk_type": "高性能",
        "price": 2835.6,
        "disk_price": 0.0,
        "pay_method": "年"
    },
    {
        "cpu": "2",
        "ram": "2",
        "room": "兆维",
        "disk_type": "高性能",
        "price": 3.2072719824,
        "disk_price": 0.0,
        "pay_method": "每天"
    },
    {
        "cpu": "2",
        "ram": "2",
        "room": "兆维",
        "disk_type": "高性能",
        "price": 136.8,
        "disk_price": 0.0,
        "pay_method": "月"
    },
    {
        "cpu": "2",
        "ram": "2",
        "room": "兆维",
        "disk_type": "高性能",
        "price": 1395.36,
        "disk_price": 0.0,
        "pay_method": "年"
    },
    {
        "cpu": "2",
        "ram": "4",
        "room": "兆维",
        "disk_type": "高性能",
        "price": 5.837882184000001,
        "disk_price": 0.0,
        "pay_method": "每天"
    },
    {
        "cpu": "2",
        "ram": "4",
        "room": "兆维",
        "disk_type": "高性能",
        "price": 207.2,
        "disk_price": 0.0,
        "pay_method": "月"
    },
    {
        "cpu": "2",
        "ram": "4",
        "room": "兆维",
        "disk_type": "高性能",
        "price": 2113.44,
        "disk_price": 0.0,
        "pay_method": "年"
    },
    {
        "cpu": "2",
        "ram": "16",
        "room": "兆维",
        "disk_type": "高性能",
        "price": 21.621543393600003,
        "disk_price": 0.0,
        "pay_method": "每天"
    },
    {
        "cpu": "2",
        "ram": "16",
        "room": "兆维",
        "disk_type": "高性能",
        "price": 591.2,
        "disk_price": 0.0,
        "pay_method": "月"
    },
    {
        "cpu": "2",
        "ram": "16",
        "room": "兆维",
        "disk_type": "高性能",
        "price": 6030.24,
        "disk_price": 0.0,
        "pay_method": "年"
    },
    {
        "cpu": "4",
        "ram": "4",
        "room": "兆维",
        "disk_type": "高性能",
        "price": 6.4145439648,
        "disk_price": 0.0,
        "pay_method": "每天"
    },
    {
        "cpu": "4",
        "ram": "4",
        "room": "兆维",
        "disk_type": "高性能",
        "price": 304.0,
        "disk_price": 0.0,
        "pay_method": "月"
    },
    {
        "cpu": "4",
        "ram": "4",
        "room": "兆维",
        "disk_type": "高性能",
        "price": 3100.8,
        "disk_price": 0.0,
        "pay_method": "年"
    },
    {
        "cpu": "4",
        "ram": "8",
        "room": "兆维",
        "disk_type": "高性能",
        "price": 11.675764368000001,
        "disk_price": 0.0,
        "pay_method": "每天"
    },
    {
        "cpu": "4",
        "ram": "8",
        "room": "兆维",
        "disk_type": "高性能",
        "price": 432.0,
        "disk_price": 0.0,
        "pay_method": "月"
    },
    {
        "cpu": "4",
        "ram": "8",
        "room": "兆维",
        "disk_type": "高性能",
        "price": 4406.4,
        "disk_price": 0.0,
        "pay_method": "年"
    },
    {
        "cpu": "4",
        "ram": "12",
        "room": "兆维",
        "disk_type": "高性能",
        "price": 16.936984771200002,
        "disk_price": 0.0,
        "pay_method": "每天"
    },
    {
        "cpu": "4",
        "ram": "12",
        "room": "兆维",
        "disk_type": "高性能",
        "price": 560.0,
        "disk_price": 0.0,
        "pay_method": "月"
    },
    {
        "cpu": "4",
        "ram": "12",
        "room": "兆维",
        "disk_type": "高性能",
        "price": 5712.0,
        "disk_price": 0.0,
        "pay_method": "年"
    },
    {
        "cpu": "4",
        "ram": "16",
        "room": "兆维",
        "disk_type": "高性能",
        "price": 22.1982051744,
        "disk_price": 0.0,
        "pay_method": "每天"
    },
    {
        "cpu": "4",
        "ram": "16",
        "room": "兆维",
        "disk_type": "高性能",
        "price": 688.0,
        "disk_price": 0.0,
        "pay_method": "月"
    },
    {
        "cpu": "4",
        "ram": "16",
        "room": "兆维",
        "disk_type": "高性能",
        "price": 7017.6,
        "disk_price": 0.0,
        "pay_method": "年"
    },
    {
        "cpu": "4",
        "ram": "32",
        "room": "兆维",
        "disk_type": "高性能",
        "price": 43.243086787200006,
        "disk_price": 0.0,
        "pay_method": "每天"
    },
    {
        "cpu": "4",
        "ram": "32",
        "room": "兆维",
        "disk_type": "高性能",
        "price": 1200.0,
        "disk_price": 0.0,
        "pay_method": "月"
    },
    {
        "cpu": "4",
        "ram": "32",
        "room": "兆维",
        "disk_type": "高性能",
        "price": 12240.0,
        "disk_price": 0.0,
        "pay_method": "年"
    },
    {
        "cpu": "8",
        "ram": "8",
        "room": "兆维",
        "disk_type": "高性能",
        "price": 12.8290879296,
        "disk_price": 0.0,
        "pay_method": "每天"
    },
    {
        "cpu": "8",
        "ram": "8",
        "room": "兆维",
        "disk_type": "高性能",
        "price": 608.0,
        "disk_price": 0.0,
        "pay_method": "月"
    },
    {
        "cpu": "8",
        "ram": "8",
        "room": "兆维",
        "disk_type": "高性能",
        "price": 6201.6,
        "disk_price": 0.0,
        "pay_method": "年"
    },
    {
        "cpu": "8",
        "ram": "16",
        "room": "兆维",
        "disk_type": "高性能",
        "price": 23.351528736000002,
        "disk_price": 0.0,
        "pay_method": "每天"
    },
    {
        "cpu": "8",
        "ram": "16",
        "room": "兆维",
        "disk_type": "高性能",
        "price": 864.0,
        "disk_price": 0.0,
        "pay_method": "月"
    },
    {
        "cpu": "8",
        "ram": "16",
        "room": "兆维",
        "disk_type": "高性能",
        "price": 8812.8,
        "disk_price": 0.0,
        "pay_method": "年"
    },
    {
        "cpu": "8",
        "ram": "24",
        "room": "兆维",
        "disk_type": "高性能",
        "price": 33.873969542400005,
        "disk_price": 0.0,
        "pay_method": "每天"
    },
    {
        "cpu": "8",
        "ram": "24",
        "room": "兆维",
        "disk_type": "高性能",
        "price": 1120.0,
        "disk_price": 0.0,
        "pay_method": "月"
    },
    {
        "cpu": "8",
        "ram": "24",
        "room": "兆维",
        "disk_type": "高性能",
        "price": 11424.0,
        "disk_price": 0.0,
        "pay_method": "年"
    },
    {
        "cpu": "8",
        "ram": "32",
        "room": "兆维",
        "disk_type": "高性能",
        "price": 44.3964103488,
        "disk_price": 0.0,
        "pay_method": "每天"
    },
    {
        "cpu": "8",
        "ram": "32",
        "room": "兆维",
        "disk_type": "高性能",
        "price": 1376.0,
        "disk_price": 0.0,
        "pay_method": "月"
    },
    {
        "cpu": "8",
        "ram": "32",
        "room": "兆维",
        "disk_type": "高性能",
        "price": 14035.2,
        "disk_price": 0.0,
        "pay_method": "年"
    },
    {
        "cpu": "8",
        "ram": "48",
        "room": "兆维",
        "disk_type": "高性能",
        "price": 65.44129196160002,
        "disk_price": 0.0,
        "pay_method": "每天"
    },
    {
        "cpu": "8",
        "ram": "48",
        "room": "兆维",
        "disk_type": "高性能",
        "price": 1888.0,
        "disk_price": 0.0,
        "pay_method": "月"
    },
    {
        "cpu": "8",
        "ram": "48",
        "room": "兆维",
        "disk_type": "高性能",
        "price": 19257.6,
        "disk_price": 0.0,
        "pay_method": "年"
    },
    {
        "cpu": "8",
        "ram": "64",
        "room": "兆维",
        "disk_type": "高性能",
        "price": 86.48617357440001,
        "disk_price": 0.0,
        "pay_method": "每天"
    },
    {
        "cpu": "8",
        "ram": "64",
        "room": "兆维",
        "disk_type": "高性能",
        "price": 2400.0,
        "disk_price": 0.0,
        "pay_method": "月"
    },
    {
        "cpu": "8",
        "ram": "64",
        "room": "兆维",
        "disk_type": "高性能",
        "price": 24480.0,
        "disk_price": 0.0,
        "pay_method": "年"
    },
    {
        "cpu": "8",
        "ram": "96",
        "room": "兆维",
        "disk_type": "高性能",
        "price": 128.57593680000002,
        "disk_price": 0.0,
        "pay_method": "每天"
    },
    {
        "cpu": "8",
        "ram": "96",
        "room": "兆维",
        "disk_type": "高性能",
        "price": 3731.2,
        "disk_price": 0.0,
        "pay_method": "月"
    },
    {
        "cpu": "8",
        "ram": "96",
        "room": "兆维",
        "disk_type": "高性能",
        "price": 38058.24,
        "disk_price": 0.0,
        "pay_method": "年"
    },
    {
        "cpu": "8",
        "ram": "128",
        "room": "兆维",
        "disk_type": "高性能",
        "price": 170.6657000256,
        "disk_price": 0.0,
        "pay_method": "每天"
    },
    {
        "cpu": "8",
        "ram": "128",
        "room": "兆维",
        "disk_type": "高性能",
        "price": 4857.6,
        "disk_price": 0.0,
        "pay_method": "月"
    },
    {
        "cpu": "8",
        "ram": "128",
        "room": "兆维",
        "disk_type": "高性能",
        "price": 49547.52,
        "disk_price": 0.0,
        "pay_method": "年"
    },
    {
        "cpu": "10",
        "ram": "16",
        "room": "兆维",
        "disk_type": "高性能",
        "price": 23.928190516800004,
        "disk_price": 0.0,
        "pay_method": "每天"
    },
    {
        "cpu": "10",
        "ram": "16",
        "room": "兆维",
        "disk_type": "高性能",
        "price": 952.0,
        "disk_price": 0.0,
        "pay_method": "月"
    },
    {
        "cpu": "10",
        "ram": "16",
        "room": "兆维",
        "disk_type": "高性能",
        "price": 9710.4,
        "disk_price": 0.0,
        "pay_method": "年"
    },
    {
        "cpu": "10",
        "ram": "24",
        "room": "兆维",
        "disk_type": "高性能",
        "price": 34.45063132320001,
        "disk_price": 0.0,
        "pay_method": "每天"
    },
    {
        "cpu": "10",
        "ram": "24",
        "room": "兆维",
        "disk_type": "高性能",
        "price": 1208.0,
        "disk_price": 0.0,
        "pay_method": "月"
    },
    {
        "cpu": "10",
        "ram": "24",
        "room": "兆维",
        "disk_type": "高性能",
        "price": 12321.6,
        "disk_price": 0.0,
        "pay_method": "年"
    },
    {
        "cpu": "10",
        "ram": "32",
        "room": "兆维",
        "disk_type": "高性能",
        "price": 44.9730721296,
        "disk_price": 0.0,
        "pay_method": "每天"
    },
    {
        "cpu": "10",
        "ram": "32",
        "room": "兆维",
        "disk_type": "高性能",
        "price": 1464.0,
        "disk_price": 0.0,
        "pay_method": "月"
    },
    {
        "cpu": "10",
        "ram": "32",
        "room": "兆维",
        "disk_type": "高性能",
        "price": 14932.8,
        "disk_price": 0.0,
        "pay_method": "年"
    },
    {
        "cpu": "10",
        "ram": "48",
        "room": "兆维",
        "disk_type": "高性能",
        "price": 66.01795374240001,
        "disk_price": 0.0,
        "pay_method": "每天"
    },
    {
        "cpu": "10",
        "ram": "48",
        "room": "兆维",
        "disk_type": "高性能",
        "price": 1976.0,
        "disk_price": 0.0,
        "pay_method": "月"
    },
    {
        "cpu": "10",
        "ram": "48",
        "room": "兆维",
        "disk_type": "高性能",
        "price": 20155.2,
        "disk_price": 0.0,
        "pay_method": "年"
    },
    {
        "cpu": "10",
        "ram": "64",
        "room": "兆维",
        "disk_type": "高性能",
        "price": 87.06283535520001,
        "disk_price": 0.0,
        "pay_method": "每天"
    },
    {
        "cpu": "10",
        "ram": "64",
        "room": "兆维",
        "disk_type": "高性能",
        "price": 2488.0,
        "disk_price": 0.0,
        "pay_method": "月"
    },
    {
        "cpu": "10",
        "ram": "64",
        "room": "兆维",
        "disk_type": "高性能",
        "price": 25377.6,
        "disk_price": 0.0,
        "pay_method": "年"
    },
    {
        "cpu": "10",
        "ram": "96",
        "room": "兆维",
        "disk_type": "高性能",
        "price": 129.1525985808,
        "disk_price": 0.0,
        "pay_method": "每天"
    },
    {
        "cpu": "10",
        "ram": "96",
        "room": "兆维",
        "disk_type": "高性能",
        "price": 3819.2,
        "disk_price": 0.0,
        "pay_method": "月"
    },
    {
        "cpu": "10",
        "ram": "96",
        "room": "兆维",
        "disk_type": "高性能",
        "price": 38955.84,
        "disk_price": 0.0,
        "pay_method": "年"
    },
    {
        "cpu": "10",
        "ram": "128",
        "room": "兆维",
        "disk_type": "高性能",
        "price": 171.2423618064,
        "disk_price": 0.0,
        "pay_method": "每天"
    },
    {
        "cpu": "10",
        "ram": "128",
        "room": "兆维",
        "disk_type": "高性能",
        "price": 4945.6,
        "disk_price": 0.0,
        "pay_method": "月"
    },
    {
        "cpu": "10",
        "ram": "128",
        "room": "兆维",
        "disk_type": "高性能",
        "price": 50445.12,
        "disk_price": 0.0,
        "pay_method": "年"
    },
    {
        "cpu": "12",
        "ram": "24",
        "room": "兆维",
        "disk_type": "高性能",
        "price": 35.02729310400001,
        "disk_price": 0.0,
        "pay_method": "每天"
    },
    {
        "cpu": "12",
        "ram": "24",
        "room": "兆维",
        "disk_type": "高性能",
        "price": 1348.8,
        "disk_price": 0.0,
        "pay_method": "月"
    },
    {
        "cpu": "12",
        "ram": "24",
        "room": "兆维",
        "disk_type": "高性能",
        "price": 13757.76,
        "disk_price": 0.0,
        "pay_method": "年"
    }
]
extend_records = [
    {
        "disk_type": "高性能",
        "pay_method": "每天",
        "extend_disk_price": 0.99,
        "extend_disk_base": 100,
        "min_extend_disk_size": 100,
        "max_extend_disk_size": 4000,
        "iops_price": 0.0,
        "min_iops_number": 0,
        "max_iops_number": 0,
        "iops_base": 0
    },
    {
        "disk_type": "高性能",
        "pay_method": "每月",
        "extend_disk_price": 50.0,
        "extend_disk_base": 100,
        "min_extend_disk_size": 100,
        "max_extend_disk_size": 4000,
        "iops_price": 0.0,
        "min_iops_number": 0,
        "max_iops_number": 0,
        "iops_base": 0
    },
    {
        "disk_type": "高性能",
        "pay_method": "每年",
        "extend_disk_price": 510.0,
        "extend_disk_base": 100,
        "min_extend_disk_size": 100,
        "max_extend_disk_size": 4000,
        "iops_price": 0.0,
        "min_iops_number": 0,
        "max_iops_number": 0,
        "iops_base": 0
    },
    {
        "disk_type": "高IO",
        "pay_method": "每天",
        "extend_disk_price": 0.71,
        "extend_disk_base": 50,
        "min_extend_disk_size": 50,
        "max_extend_disk_size": 4000,
        "iops_price": 0.07,
        "min_iops_number": 1,
        "max_iops_number": 50,
        "iops_base": 1
    },
    {
        "disk_type": "高IO",
        "pay_method": "每月",
        "extend_disk_price": 50.0,
        "extend_disk_base": 50,
        "min_extend_disk_size": 50,
        "max_extend_disk_size": 4000,
        "iops_price": 4,
        "min_iops_number": 1,
        "max_iops_number": 50,
        "iops_base": 1
    },
    {
        "disk_type": "高IO",
        "pay_method": "每年",
        "extend_disk_price": 1200.0,
        "extend_disk_base": 50,
        "min_extend_disk_size": 50,
        "max_extend_disk_size": 4000,
        "iops_price": 40.8,
        "min_iops_number": 1,
        "max_iops_number": 50,
        "iops_base": 1
    },
    {
        "disk_type": "标准型",
        "pay_method": "每天",
        "extend_disk_price": 0.99,
        "extend_disk_base": 100,
        "min_extend_disk_size": 100,
        "max_extend_disk_size": 4000,
        "iops_price": 0.0,
        "min_iops_number": 0,
        "max_iops_number": 0,
        "iops_base": 0
    },
    {
        "disk_type": "标准型",
        "pay_method": "每月",
        "extend_disk_price": 50.0,
        "extend_disk_base": 100,
        "min_extend_disk_size": 100,
        "max_extend_disk_size": 4000,
        "iops_price": 0.0,
        "min_iops_number": 0,
        "max_iops_number": 0,
        "iops_base": 0
    },
    {
        "disk_type": "标准型",
        "pay_method": "每年",
        "extend_disk_price": 510.0,
        "extend_disk_base": 100,
        "min_extend_disk_size": 100,
        "max_extend_disk_size": 4000,
        "iops_price": 0.0,
        "min_iops_number": 0,
        "max_iops_number": 0,
        "iops_base": 0
    }
]

def write_excel(header, keys, data, title, path, add_sheet=True):
    if add_sheet:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = title
    else:
        wb = openpyxl.load_workbook(path)
        wb.create_sheet(title)
        ws = wb[title]
    for field in range(1, len(header) + 1):  # 写入表头
        _ = ws.cell(row=1, column=field, value=str(header[field - 1]))

    for row1 in range(2, len(data) + 2):  # 写入数据
        for col1 in range(1, len(data[row1 - 2]) + 1):
            key = keys[col1 - 1]
            _ = ws.cell(row=row1, column=col1, value=str(data[row1 - 2][key]))

    wb.save(filename=path)


if __name__ == "__main__":
    header = ["机房", "CPU", "内存", "性能", "支付方式", "支付价格", "系统盘价格"]
    keys = ["room", "cpu", "ram", "disk_type", "pay_method", "price", "disk_price"]
    title = "CDS 价格表"
    path = "/Users/xinjie/Desktop/cds.xlsx"
    write_excel(header, keys, record, title, path)
    header = ["性能", "支付方式", "扩展盘单价", "最小拓展盘G", "最大拓展盘G", "扩展盘基数", "性能包单价", "最小性能包", "最大性能包", "性能包基数"]
    keys = ["disk_type", "pay_method", "extend_disk_price", "min_extend_disk_size", "max_extend_disk_size", "extend_disk_base", "iops_price", "min_iops_number", "max_iops_number", "iops_base"]
    title = "CDS 拓展盘价格表"
    # path = "/Users/xinjie/Desktop/cds.xlsx"
    write_excel(header, keys, extend_records, title, path, False)